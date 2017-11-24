# coding=utf-8

import sys
from openpyxl import load_workbook

'''
Student类，包含学生信息
具有属性，学号id，姓名name，成绩score
'''
class Student():
    def __init__(self):
        self.id = ''
        self.name = ''
        self.score = 0.0

'''
Data_xls类，从xls表格读取数据
属性student_list，是一个具有所有学生的信息的列表，元素是Student类
'''
class Data_xls():
    def __init__(self, file):
        # 从src加载xls表格
        wb = load_workbook(file)
        # 从表格加载第一个工作簿
        sheet = wb.worksheets[0]
        
        # 读取工作簿从第二行开始的每一行
        self.student_list = []
        for i in range(2, sheet.max_row):
            student = Student()
            student.id = str(sheet.cell(row = i, column = 1).value)                 # 第一列的信息存入student实例的id属性
            student.name = str(sheet.cell(row = i, column = 2).value)               # 第二列的信息存入student实例的name属性
            student.score = float(sheet.cell(row = i, column = 3).value)            # 第三列信息存入student实例的score属性
            self.student_list.append(student)                                       # 将一个student实例存入student_list列表中，表示一个学生的信息
            del student
        
        # 冒泡排序，得出成绩从小到大的student_list
        self.student_list_rise = self.student_list.copy()
        for i in range(len(self.student_list_rise) - 1):
            for j in range(len(self.student_list_rise) - 1 - i):
                if self.student_list_rise[j].score > self.student_list_rise[j + 1].score:
                    self.student_list_rise[j], self.student_list_rise[j + 1] = self.student_list_rise[j + 1], self.student_list_rise[j]
                    
        # 冒泡排序，得出成绩从达到小的student_list
        self.student_list_low = self.student_list.copy()
        for i in range(len(self.student_list_low) - 1):
            for j in range(len(self.student_list_low) - 1 - i):
                if self.student_list_low[j].score < self.student_list_low[j + 1].score:
                    self.student_list_low[j], self.student_list_low[j + 1] = self.student_list_low[j + 1], self.student_list_low[j]
    
    # 返回搜索结果
    def student_search(self, keyword):
        student_list_search = []
        for student in self.student_list:
            if keyword in student.name:
                student_list_search.append(student)
        return student_list_search
            
if __name__ == '__main__':
    data = Data_xls('data/33班的成绩.xlsx')
    print([student.score for student in data.student_list])
    print([student.score for student in data.student_list_rise])
    print([student.score for student in data.student_list_low])
